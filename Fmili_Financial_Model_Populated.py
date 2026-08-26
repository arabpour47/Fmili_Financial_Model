# سکریپت تزریق داده و فرمول
# pip install openpyxl
import openpyxl
from openpyxl.styles import Font

def populate_financial_model():
    # بارگذاری فایلی که در مرحله قبل ساخته شد
    wb = openpyxl.load_workbook("Fmili_Financial_Model.xlsx")
    
    # -----------------------------------------
    # 1. تکمیل صورت سود و زیان (Income Statement)
    # -----------------------------------------
    ws_is = wb["Income_Statement"]
    blue_font = Font(color="0000FF")   # فونت آبی برای داده‌های هاردکد
    black_font = Font(color="000000")  # فونت مشکی برای فرمول‌ها
    
    # ارقام تاریخی و فرضی فملی (به میلیون ریال)
    cogs_data = [102150, 188550, 310050, 384300, 517500]
    sga_data = [11350, 20950, 34450, 42700, 57500]
    tax_data = [11350, 20950, 34450, 42700, 57500] 
    
    columns = ['B', 'C', 'D', 'E', 'F'] # ستون‌های سال 1398 تا 1402
    
    for i, col in enumerate(columns):
        # بهای تمام شده (COGS)
        ws_is[f'{col}3'].value = cogs_data[i]
        ws_is[f'{col}3'].font = blue_font
        
        # فرمول سود ناخالص: Revenue - COGS
        ws_is[f'{col}4'].value = f"={col}2-{col}3"
        ws_is[f'{col}4'].font = black_font
        
        # هزینه‌های عمومی و اداری (SG&A)
        ws_is[f'{col}5'].value = sga_data[i]
        ws_is[f'{col}5'].font = blue_font
        
        # فرمول سود عملیاتی: Gross Profit - SG&A
        ws_is[f'{col}6'].value = f"={col}4-{col}5"
        ws_is[f'{col}6'].font = black_font
        
        # مالیات (Tax)
        ws_is[f'{col}7'].value = tax_data[i]
        ws_is[f'{col}7'].font = blue_font
        
        # فرمول سود خالص: Operating Income - Tax
        ws_is[f'{col}8'].value = f"={col}6-{col}7"
        ws_is[f'{col}8'].font = black_font

    # -------------------------------------------
    # 2. ساختار ترازنامه (Balance Sheet) و لینک‌های یکپارچه
    # -------------------------------------------
    ws_bs = wb["Balance_Sheet"]
    
    headers = ['Line Item', '1398', '1399', '1400', '1401', '1402 (Base)']
    for col_idx, header in enumerate(headers, 1):
        ws_bs.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
        
    bs_items = ['Cash', 'Accounts Receivable', 'Inventory', 'Total Current Assets', 
                'PPE (Net)', 'Total Assets', 'Accounts Payable', 'Short-term Debt',
                'Total Current Liabilities', 'Long-term Debt', 'Total Liabilities',
                'Paid-in Capital', 'Retained Earnings', 'Total Equity', 'Total Liab & Equity', 'Check']
                
    for row_idx, item in enumerate(bs_items, 2):
        ws_bs.cell(row=row_idx, column=1, value=item)

    # Roll-forward سود انباشته: لینک کردن سود خالص به ترازنامه
    ws_bs['B14'].value = "=Income_Statement!B8" 
    ws_bs['B14'].font = black_font
    
    for col, prev_col in zip(['C', 'D', 'E', 'F'], ['B', 'C', 'D', 'E']):
        # فرمول: سود انباشته سال قبل + سود خالص سال جاری
        ws_bs[f'{col}14'].value = f"={prev_col}14+Income_Statement!{col}8"
        ws_bs[f'{col}14'].font = black_font
        
    # فرمول کنترل ترازنامه برای سال 1402 (Check Plug)
    ws_bs['F7'].value = "=F5+F6"  # Total Assets
    ws_bs['F16'].value = "=F12+F15" # Total Liab & Equity
    ws_bs['F17'].value = "=ROUND(F7-F16, 2)" # باید صفر باشد
    ws_bs['F17'].font = black_font
    
    # ذخیره فایل تکمیل شده
    wb.save("Fmili_Financial_Model_Populated.xlsx")
    print("داده‌ها و فرمول‌ها با موفقیت تزریق شدند.")

if __name__ == "__main__":
    populate_financial_model()
